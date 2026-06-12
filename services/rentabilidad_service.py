# services/rentabilidad_service.py
import calendar
import logging
from datetime import date, datetime
from io import BytesIO

from models import get_conn

logger = logging.getLogger(__name__)

MESES_ES = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt(v):
    try:
        return f"${float(v):,.0f}".replace(',', '.')
    except Exception:
        return str(v)


def _origen_cond(origen):
    if origen == 'tiendanube':
        return "AND v.metodo_pago = 'Tienda Nube'"
    if origen == 'negocio':
        return "AND v.metodo_pago != 'Tienda Nube'"
    return ''


# ── Cálculo principal ─────────────────────────────────────────────────────────

def get_rentabilidad(fecha_desde, fecha_hasta, origen=None):
    oc = _origen_cond(origen)
    conn = get_conn()

    # Total facturado (ventas.total ya incluye descuento)
    row = conn.execute(f"""
        SELECT COALESCE(SUM(v.total), 0) AS facturado,
               COUNT(*) AS num_ventas
        FROM ventas v
        WHERE v.fecha BETWEEN ? AND ?
          AND (v.estado IS NULL OR v.estado != 'cancelada')
          {oc}
    """, (fecha_desde, fecha_hasta)).fetchone()
    facturado  = float(row['facturado'])
    num_ventas = int(row['num_ventas'])

    # Rentabilidad por producto — aplica descuento proporcional de la venta
    # CASE distribuye el descuento de la venta al item según su proporción
    prods = conn.execute(f"""
        SELECT
            p.id,
            p.descripcion,
            p.sku,
            COALESCE(p.costo, 0)  AS costo,
            SUM(dv.cantidad)       AS unidades_vendidas,
            SUM(dv.cantidad * dv.precio_unitario *
                CASE WHEN COALESCE(v.subtotal, 0) > 0
                     THEN CAST(v.total AS REAL) / v.subtotal
                     ELSE 1.0 END
            )                      AS ingresos_producto,
            AVG(dv.precio_unitario) AS precio_promedio
        FROM detalle_venta dv
        JOIN productos p ON dv.producto_id = p.id
        JOIN ventas v    ON dv.venta_id    = v.id
        WHERE v.fecha BETWEEN ? AND ?
          AND (v.estado IS NULL OR v.estado != 'cancelada')
          {oc}
        GROUP BY p.id, p.descripcion, p.sku, p.costo
    """, (fecha_desde, fecha_hasta)).fetchall()
    conn.close()

    con_costo = []
    sin_costo = []

    for r in prods:
        costo      = float(r['costo'] or 0)
        unidades   = float(r['unidades_vendidas'] or 0)
        ingresos   = float(r['ingresos_producto'] or 0)
        precio_avg = float(r['precio_promedio'] or 0)

        cogs     = costo * unidades
        ganancia = ingresos - cogs
        margen_pct  = round(ganancia / ingresos * 100, 1) if ingresos > 0 else 0
        margen_unit = precio_avg - costo

        item = {
            'id':              r['id'],
            'descripcion':     r['descripcion'],
            'sku':             r['sku'] or '',
            'costo':           costo,
            'unidades':        int(unidades),
            'precio_promedio': round(precio_avg, 2),
            'margen_unitario': round(margen_unit, 2),
            'margen_unit_pct': round(margen_unit / precio_avg * 100, 1) if precio_avg > 0 else 0,
            'ingresos':        round(ingresos, 2),
            'cogs':            round(cogs, 2),
            'ganancia':        round(ganancia, 2),
            'margen_pct':      margen_pct,
        }
        if costo > 0:
            con_costo.append(item)
        else:
            sin_costo.append(item)

    con_costo.sort(key=lambda x: x['ganancia'], reverse=True)
    sin_costo.sort(key=lambda x: x['ingresos'], reverse=True)

    total_cogs     = sum(p['cogs']    for p in con_costo)
    total_ganancia = facturado - total_cogs
    margen_bruto   = round(total_ganancia / facturado * 100, 1) if facturado > 0 else 0

    return {
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'facturado':       round(facturado, 2),
        'num_ventas':      num_ventas,
        'cogs':            round(total_cogs, 2),
        'ganancia_bruta':  round(total_ganancia, 2),
        'margen_bruto':    margen_bruto,
        'productos':       con_costo,
        'sin_costo':       sin_costo,
        'top5':            con_costo[:5],
    }


# ── Evolución margen 12 meses ─────────────────────────────────────────────────

def get_evolucion_margen_12meses(year, month, origen=None):
    oc = _origen_cond(origen)

    months = []
    for i in range(11, -1, -1):
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        months.append((y, m))

    ini_y, ini_m = months[0]
    fin_y, fin_m = months[-1]
    fecha_ini    = f"{ini_y}-{ini_m:02d}-01"
    last_day     = calendar.monthrange(fin_y, fin_m)[1]
    fecha_fin    = f"{fin_y}-{fin_m:02d}-{last_day}"

    conn = get_conn()
    ventas_rows = conn.execute(f"""
        SELECT strftime('%Y-%m', fecha) AS mes,
               COALESCE(SUM(total), 0)  AS facturado
        FROM ventas
        WHERE fecha BETWEEN ? AND ?
          AND (estado IS NULL OR estado != 'cancelada')
          {oc}
        GROUP BY mes
    """, (fecha_ini, fecha_fin)).fetchall()

    cogs_rows = conn.execute(f"""
        SELECT strftime('%Y-%m', v.fecha)       AS mes,
               SUM(dv.cantidad * COALESCE(p.costo, 0)) AS cogs
        FROM detalle_venta dv
        JOIN ventas v    ON dv.venta_id    = v.id
        JOIN productos p ON dv.producto_id = p.id
        WHERE v.fecha BETWEEN ? AND ?
          AND (v.estado IS NULL OR v.estado != 'cancelada')
          AND COALESCE(p.costo, 0) > 0
          {oc}
        GROUP BY mes
    """, (fecha_ini, fecha_fin)).fetchall()
    conn.close()

    fac_dict  = {r['mes']: float(r['facturado']) for r in ventas_rows}
    cogs_dict = {r['mes']: float(r['cogs'])      for r in cogs_rows}

    result = []
    for y, m in months:
        key      = f"{y}-{m:02d}"
        fac      = fac_dict.get(key, 0)
        cogs     = cogs_dict.get(key, 0)
        ganancia = fac - cogs
        margen   = round(ganancia / fac * 100, 1) if fac > 0 else 0
        result.append({
            'key':       key,
            'label':     f"{MESES_ES[m-1][:3]} {str(y)[2:]}",
            'facturado': fac,
            'cogs':      cogs,
            'ganancia':  ganancia,
            'margen':    margen,
        })
    return result


# ── Exportar Excel ────────────────────────────────────────────────────────────

def exportar_excel(fecha_desde, fecha_hasta, origen=None):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    data = get_rentabilidad(fecha_desde, fecha_hasta, origen)
    evol = get_evolucion_margen_12meses(
        int(fecha_hasta[:4]), int(fecha_hasta[5:7]), origen
    )

    wb        = openpyxl.Workbook()
    bold      = Font(bold=True)
    azul_font = Font(bold=True, color='FFFFFF')
    azul_fill = PatternFill('solid', fgColor='4361EE')
    gris_fill = PatternFill('solid', fgColor='F1F3F9')
    ambar_fill= PatternFill('solid', fgColor='FEF3C7')
    center    = Alignment(horizontal='center')
    money_fmt = '"$"#,##0.00'
    pct_fmt   = '0.00"%"'

    def _hrow(ws, values):
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.font = azul_font
            cell.fill = azul_fill

    def _alt(ws, start):
        for idx, row in enumerate(ws.iter_rows(min_row=start)):
            if idx % 2 == 0:
                for cell in row:
                    cell.fill = gris_fill

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen'
    ws1.merge_cells('A1:C1')
    ws1['A1'] = f"Rentabilidad — {fecha_desde} al {fecha_hasta}"
    ws1['A1'].font      = Font(bold=True, size=14, color='FFFFFF')
    ws1['A1'].fill      = azul_fill
    ws1['A1'].alignment = center
    ws1.row_dimensions[1].height = 26
    ws1.append([])
    _hrow(ws1, ['Métrica', 'Valor'])
    ws1.append(['Total facturado (con descuentos)', data['facturado']])
    ws1.append(['COGS (costo mercadería vendida)',  data['cogs']])
    ws1.append(['Ganancia bruta',                   data['ganancia_bruta']])
    ws1.append(['Margen bruto %',                   data['margen_bruto'] / 100])
    ws1.append(['Ventas incluidas',                 data['num_ventas']])
    for r in ws1.iter_rows(min_row=3, max_row=6, min_col=2, max_col=2):
        for cell in r:
            if isinstance(cell.value, float):
                cell.number_format = pct_fmt if ws1.cell(cell.row, 1).value == 'Margen bruto %' else money_fmt
    ws1['B6'].number_format = pct_fmt
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 20

    # ── Hoja 2: Por producto ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Por Producto')
    _hrow(ws2, ['Producto', 'SKU', 'Unidades', 'P. Venta Prom.',
                'Costo Unit.', 'Margen Unit.', 'Margen %',
                'Ingresos', 'COGS', 'Ganancia'])
    for p in data['productos']:
        ws2.append([
            p['descripcion'], p['sku'], p['unidades'],
            p['precio_promedio'], p['costo'],
            p['margen_unitario'], p['margen_pct'] / 100,
            p['ingresos'], p['cogs'], p['ganancia'],
        ])
    _alt(ws2, 2)
    money_cols = [4, 5, 6, 8, 9, 10]
    for col in money_cols:
        for cell in ws2.iter_rows(min_row=2, min_col=col, max_col=col):
            for c in cell:
                if isinstance(c.value, (int, float)):
                    c.number_format = money_fmt
    for cell in ws2.iter_rows(min_row=2, min_col=7, max_col=7):
        for c in cell:
            if isinstance(c.value, (int, float)):
                c.number_format = pct_fmt
    widths = [40, 14, 10, 16, 14, 14, 12, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # ── Hoja 3: Sin costo ────────────────────────────────────────────────────
    if data['sin_costo']:
        ws3 = wb.create_sheet('Sin Costo Cargado')
        _hrow(ws3, ['Producto', 'SKU', 'Unidades', 'Ingresos (sin datos de margen)'])
        for p in data['sin_costo']:
            ws3.append([p['descripcion'], p['sku'], p['unidades'], p['ingresos']])
            for cell in ws3[ws3.max_row]:
                cell.fill = ambar_fill
        for col, w in zip('ABCD', [40, 14, 10, 24]):
            ws3.column_dimensions[col].width = w

    # ── Hoja 4: Evolución margen ─────────────────────────────────────────────
    ws4 = wb.create_sheet('Evolución Margen')
    _hrow(ws4, ['Mes', 'Facturado', 'COGS', 'Ganancia', 'Margen %'])
    for e in evol:
        ws4.append([e['label'], e['facturado'], e['cogs'], e['ganancia'], e['margen'] / 100])
    _alt(ws4, 2)
    for col in ['B', 'C', 'D']:
        for row in ws4.iter_rows(min_row=2, min_col=ord(col)-64, max_col=ord(col)-64):
            for cell in row:
                cell.number_format = money_fmt
    for row in ws4.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = pct_fmt
    for col, w in zip('ABCDE', [14, 18, 18, 18, 12]):
        ws4.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Exportar PDF ──────────────────────────────────────────────────────────────

def exportar_pdf(fecha_desde, fecha_hasta, origen=None, empresa='Comenda Deco'):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    data = get_rentabilidad(fecha_desde, fecha_hasta, origen)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        title=f"Rentabilidad {fecha_desde}/{fecha_hasta} - {empresa}",
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm,  bottomMargin=2*cm
    )

    styles  = getSampleStyleSheet()
    C_AZUL  = colors.HexColor('#4361EE')
    C_GRIS  = colors.HexColor('#F1F3F9')
    C_BORDE = colors.HexColor('#DEE2E6')
    C_VERDE = colors.HexColor('#16A34A')
    C_ROJO  = colors.HexColor('#DC2626')
    C_AMBAR = colors.HexColor('#D97706')

    def _ts():
        return TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), C_AZUL),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, C_GRIS]),
            ('GRID',          (0, 0), (-1, -1), 0.5, C_BORDE),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ])

    title_st = ParagraphStyle('t', fontSize=16, textColor=C_AZUL,
                               fontName='Helvetica-Bold', spaceAfter=2)
    sub_st   = ParagraphStyle('s', fontSize=10, textColor=colors.grey, spaceAfter=14)
    sec_st   = ParagraphStyle('h', fontSize=11, textColor=C_AZUL,
                               fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=5)
    foot_st  = ParagraphStyle('f', fontSize=8, textColor=colors.grey)

    story = []
    story.append(Paragraph(empresa, title_st))
    story.append(Paragraph(
        f"Reporte de Rentabilidad — {fecha_desde} al {fecha_hasta}",
        sub_st
    ))

    # Resumen general
    story.append(Paragraph("Resumen General", sec_st))
    gbc = C_VERDE if data['ganancia_bruta'] >= 0 else C_ROJO
    sum_data = [
        ['Métrica',                         'Valor'],
        ['Total facturado (con descuentos)', _fmt(data['facturado'])],
        ['COGS (costo mercadería vendida)',   _fmt(data['cogs'])],
        ['Ganancia bruta',                   _fmt(data['ganancia_bruta'])],
        ['Margen bruto %',                   f"{data['margen_bruto']}%"],
        ['Ventas incluidas',                 str(data['num_ventas'])],
    ]
    t_sum = Table(sum_data, colWidths=[12*cm, 6*cm])
    ts_sum = _ts()
    ts_sum.add('TEXTCOLOR', (1, 3), (1, 3), gbc)
    ts_sum.add('FONTNAME',  (1, 3), (1, 4), 'Helvetica-Bold')
    t_sum.setStyle(ts_sum)
    story.append(t_sum)

    # Top productos
    if data['productos']:
        story.append(Paragraph("Top Productos por Rentabilidad", sec_st))
        top = data['productos'][:20]
        t_data = [['Producto', 'SKU', 'Unid.', 'P. Venta', 'Costo', 'Margen $', 'Margen %', 'Ganancia Total']]
        for p in top:
            t_data.append([
                p['descripcion'][:40],
                p['sku'][:14],
                str(p['unidades']),
                _fmt(p['precio_promedio']),
                _fmt(p['costo']),
                _fmt(p['margen_unitario']),
                f"{p['margen_pct']}%",
                _fmt(p['ganancia']),
            ])
        t_prod = Table(t_data, colWidths=[6.5*cm, 2.5*cm, 1.5*cm, 2.5*cm,
                                           2.5*cm, 2.5*cm, 2*cm, 3*cm])
        ts_p = _ts()
        ts_p.add('FONTSIZE', (0, 0), (-1, -1), 8)
        ts_p.add('ALIGN', (2, 0), (-1, -1), 'RIGHT')
        t_prod.setStyle(ts_p)
        story.append(t_prod)

    if data['sin_costo']:
        story.append(Spacer(1, .3*cm))
        story.append(Paragraph(
            f"⚠ {len(data['sin_costo'])} producto(s) sin costo cargado no se incluyen en el cálculo de rentabilidad.",
            ParagraphStyle('warn', fontSize=9, textColor=C_AMBAR)
        ))

    story.append(Spacer(1, .4*cm))
    story.append(Paragraph(
        "Nota: El costo refleja el valor actual del producto, no necesariamente el costo al momento de la venta.",
        ParagraphStyle('nota', fontSize=8, textColor=colors.grey)
    ))
    story.append(Spacer(1, .3*cm))
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — {empresa}",
        foot_st
    ))

    doc.build(story)
    buf.seek(0)
    return buf
