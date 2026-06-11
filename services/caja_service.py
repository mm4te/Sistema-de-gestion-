# services/caja_service.py
import io
import logging
from datetime import datetime

from models import get_conn

logger = logging.getLogger(__name__)

METODOS_CAJA = ['efectivo', 'transferencia', 'tarjeta', 'otro']
TIPOS_CAJA   = ['ingreso', 'egreso']
ORIGENES_CAJA = ['venta', 'cancelacion', 'gasto', 'manual']


def _normalizar_metodo(metodo):
    if not metodo:
        return 'otro'
    m = metodo.lower().strip()
    if m == 'efectivo':
        return 'efectivo'
    if m in ('transferencia', 'tienda nube', 'tiendanube'):
        return 'transferencia'
    if m in ('tarjeta', 'tarjeta de crédito', 'tarjeta de credito', 'débito', 'debito'):
        return 'tarjeta'
    return 'otro'


def registrar_movimiento_en_conn(conn, tipo, origen, referencia_id,
                                  descripcion, monto, metodo_pago=None,
                                  creado_por=None, fecha=None):
    """Inserta un movimiento usando una conexión existente (sin commit).
    Llamar dentro del mismo try/commit que la operación principal.
    """
    if not fecha:
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    metodo = _normalizar_metodo(metodo_pago)
    conn.execute(
        """INSERT INTO movimientos_caja
           (tipo, origen, referencia_id, descripcion, monto, metodo_pago, fecha, creado_por)
           VALUES (?,?,?,?,?,?,?,?)""",
        (tipo, origen, referencia_id, descripcion, abs(float(monto)), metodo, fecha, creado_por)
    )


def registrar_movimiento_manual(tipo, descripcion, monto, metodo_pago,
                                 creado_por, fecha=None):
    """Crea un movimiento manual con su propia conexión y commit."""
    conn = get_conn()
    try:
        if not fecha:
            fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif len(fecha) == 10:  # solo fecha 'YYYY-MM-DD'
            fecha = fecha + ' 00:00:00'
        metodo = _normalizar_metodo(metodo_pago)
        conn.execute(
            """INSERT INTO movimientos_caja
               (tipo, origen, referencia_id, descripcion, monto, metodo_pago, fecha, creado_por)
               VALUES (?,?,?,?,?,?,?,?)""",
            (tipo, 'manual', None, descripcion, abs(float(monto)), metodo, fecha, creado_por)
        )
        conn.commit()
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return True, new_id
    except Exception as e:
        conn.rollback()
        logger.error("registrar_movimiento_manual error: %s", e)
        return False, str(e)
    finally:
        conn.close()


def get_saldos(fecha_desde=None, fecha_hasta=None):
    """Devuelve saldo total y por método de pago."""
    conn = get_conn()
    conds, params = [], []
    if fecha_desde:
        conds.append("fecha >= ?");                  params.append(fecha_desde)
    if fecha_hasta:
        conds.append("fecha <= ?");                  params.append(fecha_hasta + ' 23:59:59')
    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    rows = conn.execute(f"""
        SELECT tipo, metodo_pago, SUM(monto) AS total
        FROM movimientos_caja {where}
        GROUP BY tipo, metodo_pago
    """, params).fetchall()
    conn.close()

    saldos = {'total': 0.0, 'efectivo': 0.0, 'transferencia': 0.0, 'tarjeta': 0.0, 'otro': 0.0}
    for row in rows:
        sign   = 1 if row['tipo'] == 'ingreso' else -1
        metodo = row['metodo_pago'] or 'otro'
        saldos['total'] += sign * row['total']
        if metodo in saldos:
            saldos[metodo] += sign * row['total']
        else:
            saldos['otro'] += sign * row['total']
    return saldos


def listar_movimientos(tipo=None, origen=None, metodo_pago=None,
                        fecha_desde=None, fecha_hasta=None,
                        page=1, per_page=30):
    conn = get_conn()
    conds, params = [], []
    if tipo:
        conds.append("mc.tipo = ?");         params.append(tipo)
    if origen:
        conds.append("mc.origen = ?");       params.append(origen)
    if metodo_pago:
        conds.append("mc.metodo_pago = ?");  params.append(metodo_pago)
    if fecha_desde:
        conds.append("mc.fecha >= ?");       params.append(fecha_desde)
    if fecha_hasta:
        conds.append("mc.fecha <= ?");       params.append(fecha_hasta + ' 23:59:59')
    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    total = conn.execute(
        f"SELECT COUNT(*) FROM movimientos_caja mc {where}", params
    ).fetchone()[0]

    offset = (page - 1) * per_page
    rows = conn.execute(f"""
        SELECT mc.*, u.username
        FROM movimientos_caja mc
        LEFT JOIN usuarios u ON mc.creado_por = u.id
        {where}
        ORDER BY mc.fecha DESC, mc.id DESC
        LIMIT ? OFFSET ?
    """, (*params, per_page, offset)).fetchall()
    conn.close()
    return rows, total


def exportar_excel_caja(tipo=None, origen=None, metodo_pago=None,
                         fecha_desde=None, fecha_hasta=None):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    rows, _ = listar_movimientos(tipo, origen, metodo_pago,
                                  fecha_desde, fecha_hasta, page=1, per_page=10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos de Caja"

    headers = ['ID', 'Fecha', 'Tipo', 'Origen', 'Descripción', 'Método', 'Monto', 'Usuario']
    hfill = PatternFill("solid", fgColor="4361EE")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=r['id'])
        ws.cell(row=ri, column=2, value=r['fecha'])
        ws.cell(row=ri, column=3, value=r['tipo'].capitalize())
        ws.cell(row=ri, column=4, value=r['origen'].capitalize())
        ws.cell(row=ri, column=5, value=r['descripcion'])
        ws.cell(row=ri, column=6, value=r['metodo_pago'] or '—')
        monto_cell = ws.cell(row=ri, column=7, value=r['monto'])
        monto_cell.number_format = '#,##0.00'
        monto_cell.font = Font(color="16A34A" if r['tipo'] == 'ingreso' else "DC2626")
        ws.cell(row=ri, column=8, value=r['username'] or '—')

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
