# services/resumen_service.py
import calendar
import logging
from datetime import date, datetime
from io import BytesIO

from models import get_conn

logger = logging.getLogger(__name__)

MESES_ES = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fecha_rango(year, month):
    last_day = calendar.monthrange(year, month)[1]
    return f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last_day}"


def _fmt(v):
    try:
        return f"${float(v):,.0f}".replace(',', '.')
    except Exception:
        return str(v)


# ── Datos del mes ─────────────────────────────────────────────────────────────

def get_resumen_mes(year, month):
    fecha_desde, fecha_hasta = _fecha_rango(year, month)
    conn = get_conn()

    ing = conn.execute("""
        SELECT COALESCE(SUM(total), 0) AS total, COUNT(*) AS cantidad
        FROM ventas WHERE fecha BETWEEN ? AND ? AND (estado IS NULL OR estado != 'cancelada')
    """, (fecha_desde, fecha_hasta)).fetchone()

    gas = conn.execute("""
        SELECT COALESCE(SUM(monto), 0) AS total, COUNT(*) AS cantidad
        FROM gastos WHERE fecha BETWEEN ? AND ?
    """, (fecha_desde, fecha_hasta)).fetchone()

    categorias = conn.execute("""
        SELECT cg.nombre AS categoria,
               COALESCE(SUM(g.monto), 0) AS total,
               COUNT(*) AS cantidad
        FROM gastos g
        JOIN categorias_gasto cg ON g.categoria_id = cg.id
        WHERE g.fecha BETWEEN ? AND ?
        GROUP BY cg.nombre
        ORDER BY total DESC
    """, (fecha_desde, fecha_hasta)).fetchall()

    ultimas_ventas = conn.execute("""
        SELECT v.id, v.fecha, c.nombre AS cliente, v.total, v.metodo_pago
        FROM ventas v
        JOIN clientes c ON v.cliente_id = c.id
        WHERE v.fecha BETWEEN ? AND ?
          AND (v.estado IS NULL OR v.estado != 'cancelada')
        ORDER BY v.total DESC
        LIMIT 15
    """, (fecha_desde, fecha_hasta)).fetchall()

    conn.close()

    total_ing = ing['total']
    total_gas = gas['total']
    balance   = total_ing - total_gas
    margen    = round((balance / total_ing * 100), 1) if total_ing > 0 else 0

    cat_list = []
    for c in categorias:
        pct = round((c['total'] / total_gas * 100), 1) if total_gas > 0 else 0
        cat_list.append({
            'categoria':  c['categoria'],
            'total':      c['total'],
            'cantidad':   c['cantidad'],
            'porcentaje': pct,
        })

    return {
        'year':          year,
        'month':         month,
        'mes_nombre':    MESES_ES[month - 1],
        'fecha_desde':   fecha_desde,
        'fecha_hasta':   fecha_hasta,
        'ingresos':      {'total': total_ing, 'cantidad': ing['cantidad']},
        'gastos':        {'total': total_gas, 'cantidad': gas['cantidad']},
        'balance':       balance,
        'margen':        margen,
        'por_categoria': cat_list,
        'ultimas_ventas': ultimas_ventas,
    }


# ── Evolución 12 meses ────────────────────────────────────────────────────────

def get_evolucion_12meses(year, month):
    months_list = []
    for i in range(11, -1, -1):
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        months_list.append((y, m))

    ini_y, ini_m   = months_list[0]
    fin_y, fin_m   = months_list[-1]
    fecha_desde_t  = f"{ini_y}-{ini_m:02d}-01"
    last_day_fin   = calendar.monthrange(fin_y, fin_m)[1]
    fecha_hasta_t  = f"{fin_y}-{fin_m:02d}-{last_day_fin}"

    conn = get_conn()
    ventas_rows = conn.execute("""
        SELECT strftime('%Y-%m', fecha) AS mes, COALESCE(SUM(total), 0) AS total
        FROM ventas WHERE fecha BETWEEN ? AND ?
          AND (estado IS NULL OR estado != 'cancelada')
        GROUP BY mes
    """, (fecha_desde_t, fecha_hasta_t)).fetchall()

    gastos_rows = conn.execute("""
        SELECT strftime('%Y-%m', fecha) AS mes, COALESCE(SUM(monto), 0) AS total
        FROM gastos WHERE fecha BETWEEN ? AND ?
        GROUP BY mes
    """, (fecha_desde_t, fecha_hasta_t)).fetchall()
    conn.close()

    ventas_dict = {r['mes']: r['total'] for r in ventas_rows}
    gastos_dict = {r['mes']: r['total'] for r in gastos_rows}

    result = []
    for y, m in months_list:
        key      = f"{y}-{m:02d}"
        ingresos = ventas_dict.get(key, 0)
        gastos   = gastos_dict.get(key, 0)
        result.append({
            'key':      key,
            'label':    f"{MESES_ES[m-1][:3]} {str(y)[2:]}",
            'ingresos': ingresos,
            'gastos':   gastos,
            'balance':  ingresos - gastos,
        })
    return result


# ── Exportar Excel ────────────────────────────────────────────────────────────

def exportar_excel(year, month):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    data  = get_resumen_mes(year, month)
    evol  = get_evolucion_12meses(year, month)

    wb   = openpyxl.Workbook()
    bold = Font(bold=True)
    azul_font  = Font(bold=True, color='FFFFFF')
    azul_fill  = PatternFill('solid', fgColor='4361EE')
    gris_fill  = PatternFill('solid', fgColor='F1F3F9')
    center     = Alignment(horizontal='center')

    def _header_row(ws, values):
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.font = azul_font
            cell.fill = azul_fill

    def _alt_fill(ws, start_row):
        for idx, row in enumerate(ws.iter_rows(min_row=start_row)):
            if idx % 2 == 0:
                for cell in row:
                    cell.fill = gris_fill

    # ── Resumen ──────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen'
    ws1.merge_cells('A1:C1')
    ws1['A1'] = f"Resumen Mensual — {data['mes_nombre']} {year}"
    ws1['A1'].font      = Font(bold=True, size=14, color='FFFFFF')
    ws1['A1'].fill      = azul_fill
    ws1['A1'].alignment = center
    ws1.row_dimensions[1].height = 26
    ws1.append([])

    _header_row(ws1, ['Concepto', 'Valor', 'Cantidad'])
    rows_sum = [
        ('Ingresos (ventas)', data['ingresos']['total'], data['ingresos']['cantidad']),
        ('Gastos',            data['gastos']['total'],   data['gastos']['cantidad']),
        ('Balance',           data['balance'],           ''),
        ('Margen %',          f"{data['margen']}%",      ''),
    ]
    for r in rows_sum:
        ws1.append(list(r))
    _alt_fill(ws1, 4)

    for row in ws1.iter_rows(min_row=4, max_row=6, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'

    ws1.append([])
    _header_row(ws1, ['Categoría', 'Total', 'N° Gastos', '% del Total'])
    for c in data['por_categoria']:
        ws1.append([c['categoria'], c['total'], c['cantidad'], f"{c['porcentaje']}%"])
    _alt_fill(ws1, ws1.max_row - len(data['por_categoria']) + 1)

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14

    # ── Evolución ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Evolución 12 meses')
    _header_row(ws2, ['Mes', 'Ingresos', 'Gastos', 'Balance'])
    for e in evol:
        ws2.append([e['label'], e['ingresos'], e['gastos'], e['balance']])
    _alt_fill(ws2, 2)
    for col_letter in ['B', 'C', 'D']:
        for row in ws2.iter_rows(min_row=2, min_col=ord(col_letter)-64, max_col=ord(col_letter)-64):
            for cell in row:
                cell.number_format = '"$"#,##0.00'
    ws2.column_dimensions['A'].width = 14
    for c in ['B', 'C', 'D']:
        ws2.column_dimensions[c].width = 18

    # ── Ventas del mes ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Ventas del mes')
    _header_row(ws3, ['ID', 'Fecha', 'Cliente', 'Total', 'Método de pago'])
    for v in data['ultimas_ventas']:
        ws3.append([v['id'], v['fecha'], v['cliente'], v['total'], v['metodo_pago'] or ''])
    _alt_fill(ws3, 2)
    for col, w in zip(['A','B','C','D','E'], [8, 18, 30, 14, 18]):
        ws3.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Exportar PDF ──────────────────────────────────────────────────────────────

def exportar_pdf(year, month, empresa='Comenda Deco'):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    data = get_resumen_mes(year, month)
    evol = get_evolucion_12meses(year, month)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            title=f"Resumen {month:02d}/{year} - {empresa}",
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm,  bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    C_AZUL  = colors.HexColor('#4361EE')
    C_GRIS  = colors.HexColor('#F1F3F9')
    C_BORDE = colors.HexColor('#DEE2E6')

    def _table_style(header_span=None):
        ts = [
            ('BACKGROUND',    (0, 0), (-1, 0), C_AZUL),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1,-1), 10),
            ('ROWBACKGROUNDS',(0, 1), (-1,-1), [colors.white, C_GRIS]),
            ('GRID',          (0, 0), (-1,-1), 0.5, C_BORDE),
            ('PADDING',       (0, 0), (-1,-1), 7),
        ]
        return TableStyle(ts)

    title_st = ParagraphStyle('t', fontSize=18, textColor=C_AZUL, spaceAfter=4,
                              fontName='Helvetica-Bold')
    sub_st   = ParagraphStyle('s', fontSize=11, textColor=colors.grey, spaceAfter=18)
    sec_st   = ParagraphStyle('h', fontSize=12, textColor=C_AZUL, spaceBefore=14,
                              spaceAfter=6, fontName='Helvetica-Bold')
    foot_st  = ParagraphStyle('f', fontSize=8, textColor=colors.grey)

    story = []
    story.append(Paragraph(empresa, title_st))
    story.append(Paragraph(f"Resumen Mensual — {data['mes_nombre']} {year}", sub_st))

    # Resumen
    story.append(Paragraph("Resumen General", sec_st))
    bal_color = colors.HexColor('#2ECC71') if data['balance'] >= 0 else colors.HexColor('#E63946')
    sum_data = [
        ['Concepto',          'Valor',                              'Cant.'],
        ['Ingresos (ventas)', _fmt(data['ingresos']['total']),      str(data['ingresos']['cantidad'])],
        ['Gastos',            _fmt(data['gastos']['total']),        str(data['gastos']['cantidad'])],
        ['Balance',           _fmt(data['balance']),                ''],
        ['Margen',            f"{data['margen']}%",                 ''],
    ]
    t_sum = Table(sum_data, colWidths=[8*cm, 5*cm, 4*cm])
    ts_sum = _table_style()
    ts_sum.add('TEXTCOLOR', (1, 3), (1, 3), bal_color)
    ts_sum.add('FONTNAME',  (1, 3), (1, 3), 'Helvetica-Bold')
    t_sum.setStyle(ts_sum)
    story.append(t_sum)

    # Categorías
    if data['por_categoria']:
        story.append(Paragraph("Gastos por Categoría", sec_st))
        cat_data = [['Categoría', 'Total', 'N° Gastos', '%']]
        for c in data['por_categoria']:
            cat_data.append([c['categoria'], _fmt(c['total']),
                             str(c['cantidad']), f"{c['porcentaje']}%"])
        cat_data.append(['TOTAL', _fmt(data['gastos']['total']),
                         str(data['gastos']['cantidad']), '100%'])
        t_cat = Table(cat_data, colWidths=[7*cm, 4.5*cm, 3*cm, 2.5*cm])
        ts_cat = _table_style()
        ts_cat.add('BACKGROUND', (0,-1), (-1,-1), C_GRIS)
        ts_cat.add('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold')
        t_cat.setStyle(ts_cat)
        story.append(t_cat)

    # Evolución
    story.append(Paragraph("Evolución (últimos 12 meses)", sec_st))
    evol_data = [['Mes', 'Ingresos', 'Gastos', 'Balance']]
    for e in evol:
        evol_data.append([e['label'], _fmt(e['ingresos']),
                          _fmt(e['gastos']), _fmt(e['balance'])])
    t_evol = Table(evol_data, colWidths=[3.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    t_evol.setStyle(_table_style())
    story.append(t_evol)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", foot_st
    ))

    doc.build(story)
    buf.seek(0)
    return buf
